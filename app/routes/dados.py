import io
import os
from datetime import datetime

from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required

from app.extensions import db
from app.models.categoria import Categoria
from app.models.gasto import Gasto
from app.models.gasto_fixo import GastoFixo
from app.models.instituicao import Instituicao
from app.models.investimento import Investimento
from app.models.parametro_mensal import ParametroMensal
from app.models.receita_extra import ReceitaExtra

dados_bp = Blueprint('dados', __name__, url_prefix='/dados')

MESES_PT = {
    'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
    'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
    'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12,
}

MESES_ABREV_PT = {
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12,
}

SKIP_GASTOS = {'gastos', 'total', 'salário', 'salario', 'parcial', 'sobra',
               'valor', 'valor pago', 'total geral'}

# Mapeamento de palavras-chave da descrição → categoria
KEYWORD_CAT = [
    (['visa', 'master', 'cartão', 'cartao', 'amex', 'nubank', 'c6'], 'Cartão de Crédito'),
    (['energia', 'luz', 'enel', 'cpfl', 'eletric'], 'Energia'),
    (['escola', 'faculdade', 'universidade', 'colégio', 'colegio', 'mensalidade'], 'Escola'),
    (['seguro'], 'Seguros'),
    (['netflix', 'spotify', 'amazon', 'claro', 'vivo', 'fibra', 'internet',
      'sem parar', 'condomínio', 'condominio', 'net ', 'netfibra'], 'Serviços'),
    (['previdência', 'previdencia', 'investir', 'aporte'], 'Investimentos'),
    (['alimentação', 'alimentacao', 'mercado', 'supermercado', 'ifood', 'rappi'], 'Alimentação'),
    (['uber', 'combustível', 'combustivel', 'gasolina', 'transporte'], 'Transporte'),
    (['lazer', 'viagem', 'hotel', 'cinema', 'restaurante'], 'Lazer'),
]

# Mapeamento de nome do Resumo → instituição no banco
INST_MAP = [
    (['itaú personnalite', 'itaú personnalité', 'itaú', 'itau personnalite'], 'Itaú'),
    (['xp investimentos', 'xp'], 'XP Investimentos'),
    (['ágora', 'agora', 'ágora investimentos'], 'Ágora (BTG)'),
    (['previdência bradesco', 'previdencia bradesco', 'bradesco'], 'Bradesco'),
    (['previdência itaú', 'previdencia itau', 'previdência', 'previdencia'], 'Previdência'),
    (['poupança', 'poupanca'], 'Itaú'),
]

MESES = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
         'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']


@dados_bp.route('/')
@login_required
def index():
    # Estatísticas para exibir na página
    stats = {
        'gastos': db.session.scalar(db.select(db.func.count()).select_from(Gasto)) or 0,
        'investimentos': db.session.scalar(db.select(db.func.count()).select_from(Investimento)) or 0,
        'fixos': db.session.scalar(db.select(db.func.count()).select_from(GastoFixo)) or 0,
        'meses_gastos': db.session.scalar(
            db.select(db.func.count(db.distinct(db.func.strftime('%Y-%m',
                db.func.printf('%04d-%02d', Gasto.ano, Gasto.mes)))))
        ) or 0,
    }
    db_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'data', 'banco.db'
    )
    stats['db_kb'] = round(os.path.getsize(db_path) / 1024, 1) if os.path.exists(db_path) else 0
    return render_template('dados/index.html', stats=stats)


@dados_bp.route('/backup')
@login_required
def backup():
    """Download direto do arquivo banco.db."""
    db_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'data', 'banco.db'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        db_path,
        as_attachment=True,
        download_name=f'backup_gastos_{timestamp}.db',
        mimetype='application/octet-stream',
    )


@dados_bp.route('/exportar')
@login_required
def exportar():
    """Exporta todos os dados para um arquivo Excel (.xlsx) com múltiplas abas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Helpers de estilo ──────────────────────────────────────────────────
    def header(ws, row, cols, hex_color):
        fill = PatternFill('solid', fgColor=hex_color)
        font = Font(bold=True, color='FFFFFF')
        for c, label in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=label)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws):
        for col in ws.columns:
            width = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 45)

    def brl(v):
        return float(v) if v is not None else None

    # ── Aba 1: Gastos Mensais ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Gastos Mensais'
    header(ws1, 1,
           ['Ano', 'Mês', 'Categoria', 'Tipo Cat.', 'Descrição',
            'Previsto (R$)', 'Pago (R$)', 'Diferença (R$)', 'Gerado p/ Fixo?'],
           '198754')

    gastos = db.session.scalars(
        db.select(Gasto).order_by(Gasto.ano.desc(), Gasto.mes.desc(), Gasto.categoria_id)
    ).all()
    for g in gastos:
        dif = brl(g.valor_pago) - brl(g.valor_previsto) if g.valor_pago is not None else None
        ws1.append([
            g.ano, MESES[g.mes - 1], g.categoria.nome, g.categoria.tipo,
            g.descricao, brl(g.valor_previsto), brl(g.valor_pago), dif,
            'Sim' if g.gasto_fixo_id else 'Não',
        ])
    auto_width(ws1)

    # ── Aba 2: Resumo Mensal ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Resumo Mensal')
    header(ws2, 1,
           ['Ano', 'Mês', 'Salário (R$)', 'Receitas Extras (R$)',
            'Total Entradas (R$)', 'Total Previsto (R$)',
            'Total Pago (R$)', 'Sobra (R$)'],
           '0D6EFD')

    todas_receitas = db.session.scalars(
        db.select(ReceitaExtra).order_by(ReceitaExtra.ano.desc(), ReceitaExtra.mes.desc())
    ).all()

    params = db.session.scalars(
        db.select(ParametroMensal).order_by(ParametroMensal.ano.desc(), ParametroMensal.mes.desc())
    ).all()
    for p in params:
        gm = [g for g in gastos if g.mes == p.mes and g.ano == p.ano]
        rm = [r for r in todas_receitas if r.mes == p.mes and r.ano == p.ano]
        prev = sum(brl(g.valor_previsto) for g in gm)
        pago = sum(brl(g.valor_pago) for g in gm if g.valor_pago is not None)
        extras = sum(brl(r.valor) for r in rm)
        salario = brl(p.salario)
        entradas = salario + extras
        ws2.append([p.ano, MESES[p.mes - 1], salario, extras, entradas, prev, pago, entradas - pago])
    auto_width(ws2)

    # ── Aba 3: Gastos Fixos ────────────────────────────────────────────────
    ws3 = wb.create_sheet('Gastos Fixos')
    header(ws3, 1, ['Descrição', 'Categoria', 'Valor Mensal (R$)', 'Status'], 'FD7E14')

    fixos = db.session.scalars(
        db.select(GastoFixo).order_by(GastoFixo.descricao)
    ).all()
    for f in fixos:
        ws3.append([f.descricao, f.categoria.nome, brl(f.valor), 'Ativo' if f.ativo else 'Pausado'])
    auto_width(ws3)

    # ── Aba 4: Receitas Extras ────────────────────────────────────────────
    ws4 = wb.create_sheet('Receitas Extras')
    header(ws4, 1,
           ['Ano', 'Mês', 'Tipo', 'Descrição', 'Valor (R$)', 'Instituição (se Saque)', 'Observação'],
           '20C997')

    receitas_export = db.session.scalars(
        db.select(ReceitaExtra).order_by(ReceitaExtra.ano.desc(), ReceitaExtra.mes.desc())
    ).all()
    for r in receitas_export:
        ws4.append([
            r.ano, MESES[r.mes - 1], r.tipo, r.descricao,
            brl(r.valor),
            r.instituicao.nome if r.instituicao else '',
            r.observacao or '',
        ])
    auto_width(ws4)

    # ── Aba 5: Investimentos (renumerada) ─────────────────────────────────
    ws4 = wb.create_sheet('Investimentos')
    header(ws4, 1,
           ['Ano', 'Mês', 'Corretora', 'Nome', 'Tipo', 'Emissor',
            'Risco', 'Valor (R$)', 'Vencimento', 'Fundo'],
           '6F42C1')

    investimentos = db.session.scalars(
        db.select(Investimento).order_by(
            Investimento.ano.desc(), Investimento.mes.desc(), Investimento.instituicao_id
        )
    ).all()
    for inv in investimentos:
        ws4.append([
            inv.ano, MESES[inv.mes - 1], inv.instituicao.nome,
            inv.nome, inv.tipo, inv.emissor or '',
            inv.risco, brl(inv.valor),
            inv.vencimento.strftime('%d/%m/%Y') if inv.vencimento else '',
            inv.fundo or '',
        ])
    auto_width(ws4)

    # ── Enviar arquivo ─────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        output,
        as_attachment=True,
        download_name=f'gastos_pessoais_{timestamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ── Helpers de importação ──────────────────────────────────────────────────────

def _guess_categoria(descricao: str, cats_por_nome: dict) -> Categoria | None:
    desc_lower = descricao.lower()
    for keywords, cat_nome in KEYWORD_CAT:
        if any(kw in desc_lower for kw in keywords):
            if cat_nome in cats_por_nome:
                return cats_por_nome[cat_nome]
    return None


def _guess_inst(nome: str, insts_por_nome: dict) -> Instituicao | None:
    nome_lower = nome.lower().strip()
    for keywords, inst_nome in INST_MAP:
        if any(nome_lower.startswith(kw) or kw in nome_lower for kw in keywords):
            if inst_nome in insts_por_nome:
                return insts_por_nome[inst_nome]
    return None


def _parse_gastos_sheet(ws, ano: int) -> dict:
    """Extrai gastos do Excel. Retorna {(mes, ano): [{'descricao', 'valor_previsto', 'valor_pago'}]}"""
    # Passo 1: encontrar posição das colunas de cada mês
    mes_cols = []  # [(mes_num, col_desc)]
    for row in ws.iter_rows(min_row=1, max_row=8):
        for cell in row:
            if not cell.value or not isinstance(cell.value, str):
                continue
            val = cell.value.strip().lower()
            # Verificar nome completo do mês
            for nome, num in MESES_PT.items():
                if val == nome or val.startswith(nome + '/') or val.startswith(nome + ' '):
                    if not any(m == num for m, _ in mes_cols):
                        mes_cols.append((num, cell.column))
            # Verificar abreviação (mai/26)
            for abrev, num in MESES_ABREV_PT.items():
                if val.startswith(abrev + '/') or val.startswith(abrev + ' '):
                    # Extrair ano da abreviação se presente
                    partes = val.replace('/', ' ').split()
                    if len(partes) == 2 and partes[1].isdigit():
                        a = int(partes[1])
                        ano_cell = 2000 + a if a < 100 else a
                    else:
                        ano_cell = ano
                    if not any(m == num for m, _ in mes_cols):
                        mes_cols.append((num, cell.column))

    if not mes_cols:
        return {}

    # Passo 2: para cada mês, encontrar a linha inicial de dados
    resultado = {}
    for mes_num, col_desc in mes_cols:
        # Encontrar linha do header "Gastos"
        header_row = None
        for r in ws.iter_rows(min_row=1, max_row=15, min_col=col_desc, max_col=col_desc):
            for cell in r:
                if cell.value and str(cell.value).strip().lower() == 'gastos':
                    header_row = cell.row
                    break
            if header_row:
                break

        data_start = (header_row + 1) if header_row else 4

        items = []
        for row_num in range(data_start, ws.max_row + 1):
            desc_cell = ws.cell(row=row_num, column=col_desc)
            valor_cell = ws.cell(row=row_num, column=col_desc + 1)
            pago_cell = ws.cell(row=row_num, column=col_desc + 2)

            desc = str(desc_cell.value or '').strip()
            if not desc or desc.lower() in SKIP_GASTOS:
                if desc.lower() in {'total', 'salário', 'salario'}:
                    # Tentar capturar salário
                    if desc.lower() in {'salário', 'salario'} and valor_cell.value:
                        resultado[('salario', mes_num, ano)] = float(valor_cell.value)
                    break
                continue

            valor = valor_cell.value
            pago = pago_cell.value

            if valor is None or not isinstance(valor, (int, float)):
                continue
            if float(valor) <= 0:
                continue

            items.append({
                'descricao': desc,
                'valor_previsto': float(valor),
                'valor_pago': float(pago) if isinstance(pago, (int, float)) and pago > 0 else None,
            })

        if items:
            resultado[(mes_num, ano)] = items

    return resultado


def _parse_resumo_investments(ws) -> list[dict]:
    """Lê totais de investimento por instituição na aba Resumo."""
    result = []
    in_table = False

    for row in ws.iter_rows(values_only=True):
        if not in_table:
            if row[0] and 'investimento' in str(row[0]).lower():
                in_table = True
            continue

        nome = str(row[0] or '').strip()
        if not nome:
            break
        if 'total' in nome.lower():
            break

        valor = row[1]
        if valor and isinstance(valor, (int, float)) and valor > 0:
            result.append({'nome': nome, 'valor': float(valor)})

    return result


def _guess_tipo(nome: str) -> str:
    TIPOS = [
        (['lca'], 'LCA'), (['lci'], 'LCI'), (['cdb'], 'CDB'),
        (['cri'], 'CRI'), (['coe'], 'Outros'), (['tesouro'], 'Tesouro Selic'),
        (['debênture', 'debenture', 'deb '], 'Debênture'),
        (['previdência', 'previdencia'], 'Previdência Privada'),
        (['poupança', 'poupanca'], 'Poupança'),
        (['fii'], 'FII'), (['ação', 'acoes'], 'Ação'),
    ]
    nome_lower = nome.lower()
    for kws, tipo in TIPOS:
        if any(k in nome_lower for k in kws):
            return tipo
    return 'Outros'


# ── Rota de importação ─────────────────────────────────────────────────────────

@dados_bp.route('/importar', methods=['GET', 'POST'])
@login_required
def importar():
    anos_disponiveis = list(range(datetime.now().year - 2, datetime.now().year + 2))

    if request.method == 'GET':
        return render_template('dados/importar.html', anos=anos_disponiveis)

    # ── POST: processar arquivo ────────────────────────────────────────────
    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename.endswith('.xlsx'):
        flash('Envie um arquivo .xlsx válido.', 'danger')
        return redirect(url_for('dados.importar'))

    ano = int(request.form.get('ano', datetime.now().year))
    ref_mes = int(request.form.get('ref_mes', datetime.now().month))
    ref_ano = int(request.form.get('ref_ano', datetime.now().year))
    fazer_gastos = 'fazer_gastos' in request.form
    fazer_inv = 'fazer_inv' in request.form

    try:
        import openpyxl
        wb = openpyxl.load_workbook(arquivo, data_only=True)
    except Exception as e:
        flash(f'Erro ao abrir o arquivo: {e}', 'danger')
        return redirect(url_for('dados.importar'))

    # Carregar categorias e instituições existentes
    cats_por_nome = {c.nome: c for c in db.session.scalars(db.select(Categoria)).all()}
    insts_por_nome = {i.nome: i for i in db.session.scalars(db.select(Instituicao)).all()}
    cat_padrao = next(iter(cats_por_nome.values())) if cats_por_nome else None

    resultado = {
        'gastos_importados': 0, 'gastos_pulados': 0,
        'inv_importados': 0, 'inv_pulados': 0,
        'meses': [], 'avisos': [], 'erros': [],
    }

    # ── Importar Gastos ────────────────────────────────────────────────────
    if fazer_gastos and 'Gastos' in wb.sheetnames:
        gastos_por_mes = _parse_gastos_sheet(wb['Gastos'], ano)

        for key, items in gastos_por_mes.items():
            if key[0] == 'salario':
                _, mes_num, _ = key
                salario_val = items  # type: ignore
                param = db.session.scalar(
                    db.select(ParametroMensal).where(
                        ParametroMensal.mes == mes_num, ParametroMensal.ano == ano
                    )
                )
                if not param:
                    param = ParametroMensal(mes=mes_num, ano=ano, salario=salario_val)
                    db.session.add(param)
                elif float(param.salario) == 0:
                    param.salario = salario_val
                continue

            mes_num, ano_num = key
            importados_mes = 0

            for item in items:
                # Verificar se já existe
                existe = db.session.scalar(
                    db.select(db.func.count()).select_from(Gasto).where(
                        Gasto.mes == mes_num, Gasto.ano == ano_num,
                        Gasto.descricao == item['descricao'],
                    )
                ) > 0
                if existe:
                    resultado['gastos_pulados'] += 1
                    continue

                cat = _guess_categoria(item['descricao'], cats_por_nome) or cat_padrao
                if not cat:
                    resultado['avisos'].append(f'Sem categoria disponível para "{item["descricao"]}" — pulado.')
                    resultado['gastos_pulados'] += 1
                    continue
                if cat == cat_padrao and not _guess_categoria(item['descricao'], cats_por_nome):
                    resultado['avisos'].append(
                        f'"{item["descricao"]}" sem categoria mapeada → atribuído a "{cat.nome}".'
                    )

                db.session.add(Gasto(
                    descricao=item['descricao'],
                    categoria_id=cat.id,
                    valor_previsto=item['valor_previsto'],
                    valor_pago=item['valor_pago'],
                    mes=mes_num, ano=ano_num,
                ))
                importados_mes += 1
                resultado['gastos_importados'] += 1

            if importados_mes:
                from app.routes.gastos_fixos import MESES_ABREV as MA
                resultado['meses'].append(f'{MA[mes_num-1]}/{ano_num} ({importados_mes} itens)')

        db.session.commit()

    # ── Importar Investimentos (via aba Resumo) ────────────────────────────
    if fazer_inv and 'Resumo' in wb.sheetnames:
        inv_rows = _parse_resumo_investments(wb['Resumo'])

        for row in inv_rows:
            inst = _guess_inst(row['nome'], insts_por_nome)
            if not inst:
                # Criar instituição automaticamente
                inst = Instituicao(nome=row['nome'], cor='#6c757d')
                db.session.add(inst)
                db.session.flush()
                insts_por_nome[inst.nome] = inst
                resultado['avisos'].append(f'Instituição "{row["nome"]}" criada automaticamente.')

            # Verificar duplicata
            existe = db.session.scalar(
                db.select(db.func.count()).select_from(Investimento).where(
                    Investimento.mes == ref_mes, Investimento.ano == ref_ano,
                    Investimento.nome == row['nome'],
                )
            ) > 0
            if existe:
                resultado['inv_pulados'] += 1
                continue

            db.session.add(Investimento(
                nome=row['nome'],
                tipo=_guess_tipo(row['nome']),
                instituicao_id=inst.id,
                valor=row['valor'],
                mes=ref_mes, ano=ref_ano,
                risco='baixo',
            ))
            resultado['inv_importados'] += 1

        db.session.commit()

    return render_template('dados/importar_resultado.html', r=resultado,
                           ano=ano, ref_mes=ref_mes, ref_ano=ref_ano)
